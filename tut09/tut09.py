import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def readinput():
    with open('/content/stud_list.txt', 'r') as file:
        students = [line.strip() for line in file.readlines()]
    with open('/content/dates.txt', 'r') as file:
        exec(file.read(), globals())
    attendance_data = pd.read_csv('/content/input_attendance.csv')
    return students, classes_taken_dates, class_timing, attendance_data

def processatt(students, dates, class_timing, attendance_data):
    attendance_dict = {student: {date: 0 for date in dates} for student in students}
    for _, row in attendance_data.iterrows():
        timestamp = pd.to_datetime(row['Timestamp'], dayfirst=True)
        date = timestamp.strftime("%d/%m/%Y")
        time = timestamp.strftime("%H:%M:%S")
        student_id = row['StudentID']
        if student_id in attendance_dict and date in dates:
            if class_timing[0] <= time < '19:00:00':
                attendance_dict[student_id][date] += 1
            elif '19:00:00' <= time <= class_timing[1]:
                attendance_dict[student_id][date] += 1

    return attendance_dict

def gen(attendance_dict, students, dates):
    df = pd.DataFrame(index=students, columns=dates)
    for student in students:
        for date in dates:
            df.at[student, date] = attendance_dict[student][date]
    total_count_of_dates = len(dates)
    df['Total Attendance Marked'] = df.sum(axis=1)
    df['Total Attendance allowed'] = total_count_of_dates * 2
    df['Proxy'] = abs(2*total_count_of_dates - df['Total Attendance Marked'])
    df['Total count of dates'] = total_count_of_dates
    df['Total Allowed attendance']=df.apply(lambda row: row[row< 3].sum(),axis=1)

    output_file = '/content/output_excel.xlsx'
    df.to_excel(output_file)
    coloring(output_file, df)

def coloring(file_name, df):
    wb = load_workbook(file_name)
    ws = wb.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    for row in range(2, len(df.index) + 2):
        for col in range(2, len(df.columns) + 2 - 4):  # Skip the last 4 columns (added ones)
            attendance_value = ws.cell(row=row, column=col).value
            if attendance_value == 0:
                ws.cell(row=row, column=col).fill = red_fill
            elif attendance_value == 1:
                ws.cell(row=row, column=col).fill = yellow_fill
            elif attendance_value == 2:
                ws.cell(row=row, column=col).fill = green_fill
    wb.save(file_name)

def main():
    students, dates, class_timing, attdata = readinput()
    attdict = processatt(students, dates, class_timing, attdata)
    gen(attdict, students, dates)

if __name__ == '__main__':
    main()